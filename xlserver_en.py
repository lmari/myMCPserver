from fastmcp import FastMCP
from typing import Annotated, Any
from pydantic import Field
from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries

mcp = FastMCP("Server for interaction with xlsx files", )

@mcp.tool()
def create_workbook(
        filename: Annotated[str, Field(description="Name of the xlsx file to be created")]
    ) -> dict[str, str | bool]:
    """ Create an xlsx file whose name is specified. """
    try:
        wb = Workbook()
        wb.save(filename)
        wb.close()
        return {"result": True, "message": f"File {filename} successfully created."}
    except Exception as e:
        return {"result": False, "message": f"Problems in creating file {filename}."}

@mcp.tool()
def write_data_in_cell(
        filename: Annotated[str, Field(description="Name of the xlsx file where data must be written")],
        cell: Annotated[str, Field(description="Cell where data must be written")],
        data: Annotated[Any, Field(description="Value or formula to be written in the cell")]
    ) -> dict[str, str | bool]:
    """ Assuming that the specified xlsx file already exists, write the specified value or formula in the specified cell. """
    try:
        wb = load_workbook(filename)
        ws = wb.active
        ws[cell] = data # type: ignore
        wb.save(filename)
        wb.close()
        return {"result": True, "message": f"Data successfully written in file {filename}."}
    except Exception as e:
        return {"result": False, "message": f"Problems in writing in file {filename}."}

@mcp.tool()
def write_data_in_range(
        filename: Annotated[str, Field(description="Il nome del file xlsx in cui scrivere i dati")],
        first_cell: Annotated[str, Field(description="La prima cella del range in cui scrivere")],
        last_cell: Annotated[str, Field(description="L'ultima cella del range in cui scrivere")],
        data: Annotated[list, Field(description="La lista o la lista di liste di valori o di formule da scrivere nel range")]
    ) -> dict[str, str | bool]:
    """ Assumendo che il file xlsx specificato già esista e non debba essere creato, scrive la lista specificata di valori o di formule nel range di celle specificato.
    Se il range è un vettore, i dati sono specificati come una lista; se il range è una matrice, i dati sono specificati come una lista di liste. """
    try:
        wb = load_workbook(filename)
        ws = wb.active

        min_col, min_row, max_col, max_row = range_boundaries(f"{first_cell}:{last_cell}")
        if min_col == max_col and min_row == max_row:
            ws.cell(row=min_row, column=min_col, value=data[0]) # type: ignore
        elif min_col == max_col:
            for row_idx, value in enumerate(data, start=min_row): # type: ignore
                ws.cell(row=row_idx, column=min_col, value=value) # type: ignore
        elif min_row == max_row:
            for col_idx, value in enumerate(data, start=min_col): # type: ignore
                ws.cell(row=min_row, column=col_idx, value=value) # type: ignore
        else:
            for row_idx, row_data in enumerate(data, start=min_row): # type: ignore
                for col_idx, value in enumerate(row_data, start=min_col): # type: ignore
                    ws.cell(row=row_idx, column=col_idx, value=value) # type: ignore

        wb.save(filename)
        wb.close()
        return {"result": True, "message": f"Data successfully written in file {filename}."}
    except Exception as e:
        return {"result": False, "message": f"Problems in writing in file {filename}."}



if __name__ == "__main__":
    mcp.run()

#print(write_data_in_range(**{"filename":"prova.xlsx","first_cell":"A1","last_cell":"E1","data":[1,2,3,4,5]}))