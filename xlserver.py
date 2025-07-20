from fastmcp import FastMCP
from typing import Annotated, Any
from pydantic import Field
from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries

mcp = FastMCP("Server per interazione con file xlsx")

@mcp.tool()
def create_workbook(
        filename: Annotated[str, Field(description="Il nome del file xlsx da creare")]
    ) -> dict[str, str | bool]:
    """ Crea un file xlsx dal nome specificato. """
    try:
        wb = Workbook()
        wb.save(filename)
        wb.close()
        return {"result": True, "message": f"File {filename} creato correttamente."}
    except Exception as e:
        return {"result": False, "message": f"Problemi nella creazione del file {filename}."}

@mcp.tool()
def write_data_in_cell(
        filename: Annotated[str, Field(description="Il nome del file xlsx in cui scrivere i dati")],
        cell: Annotated[str, Field(description="La cella in cui scrivere")],
        data: Annotated[Any, Field(description="Il valore o la formula da scrivere nella cella")]
    ) -> dict[str, str | bool]:
    """ Assumendo che il file xlsx specificato già esista e non debba essere creato, scrive il valore o la formula specificati nella cella specificata. """
    try:
        wb = load_workbook(filename)
        ws = wb.active
        ws[cell] = data # type: ignore
        wb.save(filename)
        wb.close()
        return {"result": True, "message": f"Dati scritti nel file {filename} correttamente."}
    except Exception as e:
        return {"result": False, "message": f"Problemi nella scrittura del file {filename}."}

@mcp.tool()
def write_data_in_range(
        filename: Annotated[str, Field(description="Il nome del file xlsx in cui scrivere i dati")],
        first_cell: Annotated[str, Field(description="La prima cella del range in cui scrivere")],
        last_cell: Annotated[str, Field(description="L'ultima cella del range in cui scrivere")],
        data: Annotated[list, Field(description="La lista o la lista di liste di valori o di formule da scrivere nel range")]
    ) -> dict[str, str | bool]:
    """ Assumendo che il file xlsx specificato già esista e non debba essere creato, scrive la lista specificata di valori o di formule nel range di celle specificato.
    Se il range è un vettore, i dati sono specificati come una lista; se il range è una matrice, i dati sono specificati come una lista di liste. """
    try:
        wb = load_workbook(filename)
        ws = wb.active

        min_col, min_row, max_col, max_row = range_boundaries(f"{first_cell}:{last_cell}")
        if min_col == max_col and min_row == max_row:
            ws.cell(row=min_row, column=min_col, value=data[0]) # type: ignore
        elif min_col == max_col:
            for row_idx, value in enumerate(data, start=min_row): # type: ignore
                ws.cell(row=row_idx, column=min_col, value=value) # type: ignore
        elif min_row == max_row:
            for col_idx, value in enumerate(data, start=min_col): # type: ignore
                ws.cell(row=min_row, column=col_idx, value=value) # type: ignore
        else:
            for row_idx, row_data in enumerate(data, start=min_row): # type: ignore
                for col_idx, value in enumerate(row_data, start=min_col): # type: ignore
                    ws.cell(row=row_idx, column=col_idx, value=value) # type: ignore

        wb.save(filename)
        wb.close()
        return {"result": True, "message": f"Dati scritti nel file {filename} correttamente."}
    except Exception as e:
        return {"result": False, "message": f"Problemi nella scrittura del file {filename}."}

import sys
if __name__ == "__main__":
    msg = """Attiva il server specificando il parametro 'stdio' per esecuzione locale,
oppure 'http' e in questo caso anche l'indirizzo IP e la porta TCP,
dunque per esempio 'http 127.0.0.1 8000'"""
    if len(sys.argv) == 1:
        print(msg)
        sys.exit(1)
    if sys.argv[1] == "stdio":
        mcp.run(transport="stdio")
    elif sys.argv[1] == "http":
        mcp.run(transport="streamable-http", host=sys.argv[2], port=int(sys.argv[3]),
                path="/mcp", log_level="debug")
    else:
        print(msg)
